from openpyxl import load_workbook, Workbook
import numpy as np
from datetime import datetime, timedelta
import netCDF4

load_xls=False
if load_xls:
    xlsfile="/Users/gbolzon/Downloads/DATA_ABS_QC_BOUSSOLE_Update_10sept2018.xlsx"
    
    wb = load_workbook(filename=xlsfile, read_only=False,data_only=True)
    
    sheet = "Aphy"
    print "reading " + sheet
    ws=wb[sheet]

col_day   = 3
col_month = 4
col_year  = 5
col_time  = 6
col_depth = 9
col_chla = 10


lambda_start=300
lambda_end  =794
nlambda = lambda_end - lambda_start + 1
#my_dtype=[('time',np.float32),('depth',np.float32),('chla',np.float32)]
#for Lambda in range(lambda_start,lambda_end+1):
#    my_dtype.append(  ("l%03d"%Lambda, np.float32) )

nP = 3502
Dref = datetime(1970,1,1,0,0,0)

TIMES = np.zeros((nP,),np.float64)
DEPTH = np.zeros((nP,),np.float32)
CHLA  = np.zeros((nP,),np.float32)

LAMBDA = np.zeros((nP,nlambda), np.float32)

for i in range(nP):
    row=i+3
    year  = int(ws.cell(row=row, column = col_year).value)
    month = int(ws.cell(row=row, column = col_month).value)
    day   = int(ws.cell(row=row, column = col_day).value)
    hr    = float(ws.cell(row=row, column = col_time).value)
    d = datetime(year,month,day) + timedelta(hours = hr)
    Diff = d-Dref
    TIMES[i] = Diff.total_seconds()
    DEPTH[i] = float(ws.cell(row=row, column = col_depth).value)
    chla = float(ws.cell(row=row, column = col_chla).value)
    if chla > 90:
        CHLA[i] = chla/1000.
    else:
        CHLA[i] = chla
    for ilambda in range(nlambda):
        value = ws.cell(row=row, column = ilambda+11).value
        if str(value)=='NA':
            LAMBDA[i,ilambda] = np.nan
        else:
            LAMBDA[i,ilambda] = float(value)
    

unique_times, indices_r = np.unique(TIMES,  return_inverse=True)

output_records=0
for it, t in enumerate(unique_times):
    ii_same_time = indices_r==it
    unique_depth = np.unique(DEPTH[ii_same_time])
    output_records += unique_depth.size



TIMES_out = np.zeros((output_records,),np.float32)
DEPTH_out = np.zeros((output_records,),np.float32)
CHLA_out  = np.zeros((output_records,),np.float32)

LAMBDA_out = np.zeros((output_records,nlambda), np.float32)


j_out = 0
for it, t in enumerate(unique_times):
    ii_same_time = indices_r==it

    indexes_time = np.nonzero(ii_same_time)[0]
    unique_depth, depth_r_indexes, counts = np.unique(DEPTH[ii_same_time], return_inverse=True, return_counts=True)
    
    for idepth, depth in enumerate(unique_depth):
        if counts[idepth] > 1:
            ii_same_depth = depth_r_indexes==idepth
            indexes_multiples = indexes_time[ii_same_depth]

            TIMES_out[j_out] = TIMES[indexes_multiples[0]]
            DEPTH_out[j_out] = DEPTH[indexes_multiples[0]]
            LAMBDA_out[j_out,:] = LAMBDA[indexes_multiples,:].mean(axis=0)
            CHLA_out[ j_out] =   CHLA[indexes_multiples].mean()
        else:
            j_in = indexes_time[depth_r_indexes==idepth][0]
            TIMES_out[j_out] = TIMES[j_in]
            DEPTH_out[j_out] = DEPTH[j_in]
            CHLA_out[j_out]  = CHLA[j_in]
            LAMBDA_out[j_out,:] = LAMBDA[j_in,:]
        j_out +=1



ncOUT = netCDF4.Dataset('out.nc','w')
ncOUT.createDimension('n',output_records)
ncOUT.createDimension('lambda', nlambda)


ncvar=ncOUT.createVariable('time','d',('n',))
setattr(ncvar,'units',       'seconds since 1970-01-01 00:00:00')
setattr(ncvar,'long_name'    ,'time')
setattr(ncvar,'standard_name','time')
setattr(ncvar,'axis'         ,'T')
setattr(ncvar,'calendar'     ,'standard')
ncvar[:] = TIMES_out
ncvar=ncOUT.createVariable('chla','f',('n',))
ncvar[:] = CHLA_out
ncvar=ncOUT.createVariable('depth','f',('n',))
setattr(ncvar,'units'        ,'m')
setattr(ncvar,'long_name'    ,'depth')
setattr(ncvar,'standard_name','depth')
setattr(ncvar,'positive'     ,'down')
setattr(ncvar,'axis'         ,'Z')
ncvar[:] = DEPTH_out
ncvar = ncOUT.createVariable('lambda','f',('n','lambda'))
ncvar[:] = LAMBDA_out
ncOUT.close()